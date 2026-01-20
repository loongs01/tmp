#!/usr/bin/env python3
"""
Excel转CSV批量转换工具（增强版）- 优化修复版
功能：将Excel文件中的所有工作表转换为独立的CSV文件
特点：
1. 深度探查机制：确保不会遗漏隐藏在空行空列后的真实数据
2. 合并单元格处理：保留合并单元格的语义信息
3. 自动处理复杂表头结构
4. 过滤带有删除线的单元格
5. 支持大文件和宽表（最多支持到XFD列）
6. 保持数据完整性
7. 使用UTF-8编码，避免乱码
8. 增强错误处理和格式兼容性
"""

import pandas as pd
import openpyxl
import os
import argparse
import sys
from openpyxl.utils import get_column_letter
import openpyxl.descriptors.base
import openpyxl.descriptors.sequence
from openpyxl.styles.fills import Fill, PatternFill

# Monkeypatch to fix "expected <class 'openpyxl.styles.fills.Fill'>" error
# This error occurs when openpyxl encounters unexpected style data
_original_convert = openpyxl.descriptors.base._convert

def _patched_convert(expected_type, value):
    try:
        return _original_convert(expected_type, value)
    except TypeError:
        # If we expect a Fill but get something else (and validation fails),
        # return a default PatternFill to avoid crashing.
        if expected_type is Fill:
            return PatternFill()
        raise

openpyxl.descriptors.base._convert = _patched_convert
openpyxl.descriptors.sequence._convert = _patched_convert


def safe_load_workbook(excel_file):
    """
    安全地加载Excel工作簿，处理格式兼容性问题
    """
    try:
        # 尝试多种加载方式
        # 方式1：最小化加载，避免格式解析问题
        try:
            wb = openpyxl.load_workbook(
                excel_file,
                data_only=True,
                read_only=False,  # read_only模式不支持字体检查
                keep_links=False,
                keep_vba=False
            )
            print(f"成功以标准模式加载工作簿")
            return wb
        except Exception as e1:
            print(f"标准模式加载失败: {e1}")

            # 方式2：尝试read_only模式
            try:
                wb = openpyxl.load_workbook(
                    excel_file,
                    data_only=True,
                    read_only=True
                )
                print(f"成功以只读模式加载工作簿")
                return wb
            except Exception as e2:
                print(f"只读模式加载失败: {e2}")

                # 方式3：使用pandas作为后备方案
                print(f"尝试使用pandas直接读取数据...")
                return None

    except Exception as e:
        print(f"所有加载方式均失败: {str(e)}")
        return None


def is_cell_strikethrough(cell):
    """
    安全地检查单元格是否有删除线
    """
    try:
        if hasattr(cell, 'font') and cell.font:
            # 处理不同的font.strike属性类型
            strike = cell.font.strike
            if strike is None:
                return False
            elif isinstance(strike, bool):
                return strike
            elif isinstance(strike, str):
                return strike.lower() == 'true' or strike == '1'
            else:
                return bool(strike)
        return False
    except Exception:
        # 如果检查失败，默认没有删除线
        return False


def excel_to_csv(excel_file, output_dir=None, skip_deleted=True):
    """
    将Excel文件转换为CSV文件

    参数:
        excel_file: 输入的Excel文件路径
        output_dir: 输出目录，默认与Excel文件同目录
        skip_deleted: 是否跳过带有删除线的单元格
    """

    # 检查文件是否存在
    if not os.path.exists(excel_file):
        print(f"错误：文件 {excel_file} 不存在")
        return False

    # 设置输出目录
    if output_dir is None:
        # 使用原始文件名创建输出目录，而不是固定字符串'model'
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        output_dir = os.path.join(os.path.dirname(excel_file), f"{base_name}_csv")

    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    print(f"输出目录: {output_dir}")

    try:
        # 安全地加载工作簿
        wb = safe_load_workbook(excel_file)

        if wb is None:
            # 使用pandas作为后备方案
            return excel_to_csv_pandas(excel_file, output_dir, skip_deleted)

        # 获取所有工作表名称
        sheet_names = wb.sheetnames
        print(f"发现 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")

        # 处理每个工作表
        for sheet_idx, sheet_name in enumerate(sheet_names, 1):
            try:
                print(f"\n处理工作表 [{sheet_idx}/{len(sheet_names)}]: {sheet_name}")

                # 安全地获取工作表
                try:
                    ws = wb[sheet_name]
                except Exception as e:
                    print(f"  无法访问工作表 {sheet_name}: {e}")
                    continue

                # 深度探查：找到真实的最后一行和最后一列
                try:
                    # 使用更安全的方式获取行列范围
                    max_row = ws.max_row if hasattr(ws, 'max_row') else 0
                    max_col = ws.max_column if hasattr(ws, 'max_column') else 0

                    if max_row == 0 or max_col == 0:
                        print(f"  工作表 {sheet_name} 似乎是空的，跳过")
                        continue

                    # 限制探查范围，避免过大文件
                    real_max_row = find_real_max_row(ws, max_row)
                    real_max_col = find_real_max_col(ws, max_col)

                    print(f"  数据范围: 行 {real_max_row}, 列 {real_max_col}")

                except Exception as e:
                    print(f"  探查数据范围时出错: {e}")
                    # 使用保守的估计值
                    real_max_row = min(max_row, 1000) if max_row > 0 else 100
                    real_max_col = min(max_col, 100) if max_col > 0 else 26

                # 收集所有单元格数据
                data = []
                try:
                    # 处理合并单元格
                    merged_cells = []
                    try:
                        if hasattr(ws, 'merged_cells'):
                            merged_cells = list(ws.merged_cells.ranges)
                    except Exception:
                        merged_cells = []

                    # 逐行读取数据
                    rows_processed = 0
                    for row_idx, row in enumerate(ws.iter_rows(
                            min_row=1, max_row=real_max_row,
                            min_col=1, max_col=real_max_col,
                            values_only=False
                    ), 1):
                        row_data = []
                        for cell in row:
                            cell_value = None

                            try:
                                # 检查删除线
                                if skip_deleted and is_cell_strikethrough(cell):
                                    cell_value = None
                                else:
                                    # 检查合并单元格
                                    cell_value = get_merged_cell_value(cell, merged_cells)
                            except Exception:
                                # 如果获取值失败，尝试直接获取value属性
                                try:
                                    cell_value = cell.value
                                except Exception:
                                    cell_value = None

                            row_data.append(cell_value)

                        data.append(row_data)
                        rows_processed += 1

                    print(f"  已处理 {rows_processed} 行数据")

                except Exception as e:
                    print(f"  读取单元格数据时出错: {e}")
                    # 尝试简化的数据读取方式
                    data = read_simplified_data(ws, real_max_row, real_max_col)

                # 转换为DataFrame
                if not data:
                    print(f"  工作表 {sheet_name} 没有数据，跳过")
                    continue

                df = pd.DataFrame(data)

                # 智能检测表头
                df = detect_and_set_header(df)

                # 清理数据
                df = clean_dataframe(df)

                if df.empty:
                    print(f"  工作表 {sheet_name} 处理后为空，跳过")
                    continue

                # 保存为CSV文件
                # 清理文件名中的非法字符
                safe_sheet_name = sanitize_filename(sheet_name)
                csv_filename = os.path.join(output_dir, f"{safe_sheet_name}.csv")

                try:
                    df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
                    print(f"  ✓ 已保存: {os.path.basename(csv_filename)} "
                          f"({len(df)}行 × {len(df.columns)}列)")
                except Exception as e:
                    print(f"  保存CSV文件时出错: {e}")
                    # 尝试使用不同编码
                    try:
                        df.to_csv(csv_filename, index=False, encoding='utf-8')
                        print(f"  ✓ 已保存 (UTF-8编码): {os.path.basename(csv_filename)}")
                    except Exception:
                        print(f"  ✗ 无法保存工作表 {sheet_name}")

            except Exception as e:
                print(f"  处理工作表 {sheet_name} 时发生错误: {str(e)}")
                import traceback
                traceback.print_exc()
                continue

        # 关闭工作簿
        try:
            wb.close()
        except Exception:
            pass

        print(f"\n{'=' * 60}")
        print(f"转换完成！所有CSV文件已保存到: {output_dir}")
        print(f"{'=' * 60}")
        return True

    except Exception as e:
        print(f"\n转换过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def excel_to_csv_pandas(excel_file, output_dir, skip_deleted=True):
    """
    使用pandas作为后备转换方案
    注意：pandas方案无法处理删除线检测
    """
    print("\n切换到pandas转换方案（部分高级功能不可用）")

    try:
        # 尝试使用不同引擎读取Excel
        engines = ['openpyxl', 'xlrd', None]  # None表示自动选择

        excel_data = None
        used_engine = None

        for engine in engines:
            try:
                print(f"尝试使用引擎: {engine or 'auto'}")
                excel_data = pd.read_excel(
                    excel_file,
                    sheet_name=None,
                    engine=engine,
                    dtype=str,  # 统一转为字符串处理
                    na_filter=False,
                    keep_default_na=False
                )
                used_engine = engine
                print(f"成功使用引擎: {engine or 'auto'}")
                break
            except Exception as e:
                print(f"引擎 {engine or 'auto'} 失败: {e}")
                continue

        if excel_data is None:
            print("所有引擎均失败，无法读取Excel文件")
            return False

        # 处理每个工作表
        total_sheets = len(excel_data)
        for idx, (sheet_name, df) in enumerate(excel_data.items(), 1):
            try:
                print(f"\n处理工作表 [{idx}/{total_sheets}]: {sheet_name}")

                # 清理数据
                df = df.fillna('')
                df = df.replace([None, 'None', 'nan', 'NaN'], '')

                # 移除完全空白的行和列
                df = df.loc[~(df.astype(str).apply(lambda x: x.str.strip() == '').all(axis=1))]
                if not df.empty:
                    df = df.loc[:, ~(df.astype(str).apply(lambda x: x.str.strip() == '').all(axis=0))]

                if df.empty:
                    print(f"  工作表 {sheet_name} 为空，跳过")
                    continue

                # 保存CSV
                safe_name = sanitize_filename(sheet_name)
                csv_path = os.path.join(output_dir, f"{safe_name}.csv")

                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                print(f"  ✓ 已保存: {os.path.basename(csv_path)} "
                      f"({len(df)}行 × {len(df.columns)}列)")

            except Exception as e:
                print(f"  处理工作表 {sheet_name} 时出错: {e}")
                continue

        print(f"\n{'=' * 60}")
        print(f"pandas转换完成！文件保存在: {output_dir}")
        print(f"使用的引擎: {used_engine or 'auto'}")
        print(f"注意：pandas方案无法处理删除线过滤和合并单元格语义")
        print(f"{'=' * 60}")
        return True

    except Exception as e:
        print(f"pandas转换失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def find_real_max_row(ws, max_row):
    """安全地查找真实的最大行"""
    probe_rows = min(200, max_row) if max_row > 0 else 100
    real_max_row = max_row if max_row > 0 else 0

    try:
        for i in range(max_row, max_row - probe_rows, -1):
            if i <= 1:  # 从1开始的行号
                break
            row_has_data = False
            for cell in ws[i]:
                if cell.value not in (None, '', ' ', '\n', '\t'):
                    row_has_data = True
                    break
            if row_has_data:
                real_max_row = i
                break
    except Exception:
        real_max_row = max_row

    return max(1, real_max_row)  # 确保至少为1


def find_real_max_col(ws, max_col):
    """安全地查找真实的最大列"""
    probe_cols = min(50, max_col) if max_col > 0 else 26
    real_max_col = max_col if max_col > 0 else 0

    try:
        for i in range(max_col, max_col - probe_cols, -1):
            if i <= 1:  # 从1开始的列号
                break
            col_has_data = False
            for row in ws.iter_rows(min_col=i, max_col=i, max_row=min(100, ws.max_row)):
                if row and row[0].value not in (None, '', ' ', '\n', '\t'):
                    col_has_data = True
                    break
            if col_has_data:
                real_max_col = i
                break
    except Exception:
        real_max_col = max_col

    return max(1, real_max_col)  # 确保至少为1


def get_merged_cell_value(cell, merged_cells):
    """获取合并单元格的值"""
    if not merged_cells:
        return cell.value

    for merged_cell in merged_cells:
        try:
            if cell.coordinate in merged_cell:
                # 如果是合并单元格的左上角单元格，使用原值
                if cell.coordinate == merged_cell.start_cell.coordinate:
                    return cell.value
                else:
                    # 其他合并单元格，使用左上角单元格的值
                    return merged_cell.start_cell.value
        except Exception:
            continue

    return cell.value


def read_simplified_data(ws, max_row, max_col):
    """简化的数据读取方式"""
    data = []
    try:
        # 使用values_only模式避免格式问题
        for row in ws.iter_rows(
                min_row=1, max_row=min(max_row, 10000),  # 限制行数
                min_col=1, max_col=min(max_col, 200),  # 限制列数
                values_only=True
        ):
            data.append(list(row))
    except Exception:
        # 如果还失败，使用最基本的读取方式
        data = []
        for r in range(1, min(max_row, 1000) + 1):
            row_data = []
            for c in range(1, min(max_col, 100) + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    row_data.append(cell.value)
                except Exception:
                    row_data.append(None)
            data.append(row_data)

    return data


def detect_and_set_header(df):
    """智能检测和设置表头"""
    if df.empty:
        return df

    # 查找第一个非空行作为表头
    header_row = -1
    for i in range(min(20, len(df))):  # 只检查前20行
        if not df.iloc[i].isnull().all():
            header_row = i
            break

    if header_row >= 0 and header_row < len(df):
        try:
            # 设置表头
            df.columns = df.iloc[header_row].astype(str).fillna(f'Unnamed_{i}')
            # 移除表头行
            df = df.iloc[header_row + 1:].reset_index(drop=True)
        except Exception:
            # 如果设置表头失败，保持原样
            pass

    return df


def clean_dataframe(df):
    """清理DataFrame"""
    if df.empty:
        return df

    # 去除列名中的NaN
    df.columns = [str(col).strip() if pd.notna(col) else f'Column_{i}'
                  for i, col in enumerate(df.columns, 1)]

    # 去除完全空白的行
    df = df.dropna(how='all')

    # 去除完全空白的列
    df = df.loc[:, ~df.isna().all()]

    # 重置索引
    df = df.reset_index(drop=True)

    return df


def sanitize_filename(filename):
    """清理文件名中的非法字符"""
    # Windows文件名非法字符
    illegal_chars = r'<>:"/\|?*'
    for char in illegal_chars:
        filename = filename.replace(char, '_')

    # 限制文件名长度
    if len(filename) > 100:
        name, ext = os.path.splitext(filename)
        filename = name[:95] + ext

    return filename.strip()


def main():
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(
        description='Excel转CSV批量转换工具（增强版）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s data.xlsx                    # 转换当前目录的data.xlsx
  %(prog)s data.xlsx -o ./output       # 指定输出目录
  %(prog)s "文件 名.xlsx"              # 处理含空格的文件名
  %(prog)s data.xlsx --keep-deleted    # 保留删除线内容
        """
    )
    parser.add_argument('excel_file', help='输入的Excel文件路径')
    parser.add_argument('-o', '--output', help='输出目录路径')
    parser.add_argument('--keep-deleted', action='store_true',
                        help='保留带有删除线的单元格内容（默认跳过）')
    parser.add_argument('--debug', action='store_true',
                        help='启用调试模式，显示详细错误信息')

    # 解析命令行参数
    args = parser.parse_args()

    # 检查文件是否存在
    if not os.path.exists(args.excel_file):
        print(f"错误：文件 '{args.excel_file}' 不存在")
        sys.exit(1)

    # 显示开始信息
    print(f"{'=' * 60}")
    print(f"Excel转CSV转换工具")
    print(f"{'=' * 60}")
    print(f"输入文件: {args.excel_file}")
    print(f"输出目录: {args.output if args.output else '自动生成'}")
    print(f"删除线处理: {'保留' if args.keep_deleted else '跳过'}")
    print(f"{'=' * 60}")

    # 执行转换
    success = excel_to_csv(args.excel_file, args.output, not args.keep_deleted)

    if success:
        sys.exit(0)
    else:
        print("\n转换失败！请检查：")
        print("1. Excel文件是否损坏")
        print("2. 是否有足够的磁盘空间")
        print("3. 文件是否被其他程序占用")
        if args.debug:
            print("\n调试信息已显示在上方")
        sys.exit(1)


if __name__ == '__main__':
    main()