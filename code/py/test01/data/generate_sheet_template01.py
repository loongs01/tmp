from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def create_metric_template(save_path=None):
    """生成包含用户多维度指标的Excel模板"""
    wb = Workbook()

    # ===== 1. 封面页 =====
    ws_cover = wb.active
    ws_cover.title = "封面"
    cover_data = [
        ["用户多维指标口径文档"],
        ["项目名称", "经营分析"],
        ["版本号", "V1.0"],
        ["编制部门", "需求部门"],
        ["负责人", "xxx"],
        ["联系方式", "xxx"],
        ["更新日期", datetime.now().strftime("%Y-%m-%d")],
        ["说明", "此文档定义用户特征、健康、职业等维度的指标口径"]
    ]
    for row in cover_data:
        ws_cover.append(row)

    # ===== 2. 目录页 =====
    ws_toc = wb.create_sheet("目录")
    toc_data = [
        ["序号", "工作表名称", "内容说明", "最后更新日期"],
        [1, "指标总览", "核心指标清单", datetime.now().strftime("%Y-%m-%d")],
        [2, "基础特征指标", "用户画像基础指标", datetime.now().strftime("%Y-%m-%d")],
        [3, "健康指标", "健康体征与行为指标", datetime.now().strftime("%Y-%m-%d")],
        [4, "职业发展指标", "职业与教育指标", datetime.now().strftime("%Y-%m-%d")],
        [5, "财务指标", "资产与消费指标", datetime.now().strftime("%Y-%m-%d")],
        [6, "行为指标", "休闲娱乐行为指标", datetime.now().strftime("%Y-%m-%d")],
        [7, "复合指标", "衍生计算指标", datetime.now().strftime("%Y-%m-%d")],
        [8, "变更记录", "版本修订历史", datetime.now().strftime("%Y-%m-%d")]
    ]
    for row in toc_data:
        ws_toc.append(row)

    # ===== 3. 指标总览 =====
    ws_summary = wb.create_sheet("指标总览")
    summary_headers = [
        "指标ID", "指标名称", "业务域", "优先级",
        "数据来源表", "更新频率", "负责人"
    ]
    ws_summary.append(summary_headers)

    summary_data = [
        ["USER_001", "高价值用户标识", "用户特征", "P0", "dws_user_characteristic_aggregate_di", "每日", "xxx"],
        ["HEALTH_002", "BMI指数", "健康", "P1", "dws_user_health_summary_di", "每日", "xxx"],
        ["CAREER_003", "职业满意度", "职业", "P1", "dws_user_career_info", "每月", "xxx"],
        ["FIN_004", "净资产", "财务", "P1", "dws_user_financial_management_di", "季度", "xxx"],
        ["BEHAV_005", "日均休闲时长", "行为", "P2", "dws_user_leisure_entertainment_di", "每周", "xxx"]
    ]
    for row in summary_data:
        ws_summary.append(row)

    # ===== 4. 基础特征指标 =====
    ws_basic = wb.create_sheet("基础特征指标")
    basic_headers = [
        "指标ID", "指标名称", "业务定义", "计算公式/规则",
        "数据字段", "统计口径", "特殊说明"
    ]
    ws_basic.append(basic_headers)

    basic_data = [
        ["USER_101", "用户生命周期", "用户当前所处阶段",
         "CASE WHEN last_active_date<30d THEN '流失' ELSE...",
         "user_lifecycle", "每日计算", "基于RFM模型"],
        ["USER_102", "RFM评分", "用户价值分层",
         "CONCAT(recency_score,'-',frequency_score,'-',monetary_score)",
         "rfm_score", "月度更新", "需排除测试账号"],
        ["USER_103", "内容创作者标识", "是否UGC生产者",
         "is_content_creator=1 AND content_count>5",
         "is_content_creator", "每日", "需人工审核"]
    ]
    for row in basic_data:
        ws_basic.append(row)

    # ===== 5. 健康指标 =====
    ws_health = wb.create_sheet("健康指标")
    health_headers = [
        "指标ID", "指标名称", "医学定义", "计算公式/规则",
        "正常范围", "风险等级", "数据来源"
    ]
    ws_health.append(health_headers)

    health_data = [
        ["HLTH_201", "BMI指数", "体重(kg)/身高(m)^2",
         "weight/(height/100)^2",
         "18.5-24", "1级风险>30", "健康表height/weight"],
        ["HLTH_202", "睡眠质量", "睡眠质量综合评分",
         "(deep_sleep_ratio*0.6)+(sleep_duration*0.4)",
         "80-100分", "2级风险<60", "手环数据"],
        ["HLTH_203", "高血压风险", "血压持续偏高",
         "avg_systolic>140 OR avg_diastolic>90",
         "<140/90", "3级风险", "医疗记录"]
    ]
    for row in health_data:
        ws_health.append(row)

    # ===== 6. 职业发展指标 =====
    ws_career = wb.create_sheet("职业发展指标")
    career_headers = [
        "指标ID", "指标名称", "定义", "计算公式/规则",
        "分级标准", "分析维度", "数据表"
    ]
    ws_career.append(career_headers)

    career_data = [
        ["CR_301", "职业满意度", "对当前工作满意程度",
         "career_satisfaction_score",
         "1-3分:低 4-7:中 8-10:高", "行业/职级", "dws_user_career_info"],
        ["CR_302", "薪资水平", "当前月收入",
         "current_salary",
         "按行业百分位", "地区/经验", "职业表"],
        ["CR_303", "技能水平", "编程能力等级",
         "skill_level_coding",
         "1-初级 2-中级 3-高级", "教育背景", "教育表"]
    ]
    for row in career_data:
        ws_career.append(row)

    # ===== 7. 财务指标 =====
    ws_finance = wb.create_sheet("财务指标")
    finance_headers = [
        "指标ID", "指标名称", "财务定义", "计算公式/规则",
        "健康阈值", "分析建议", "更新周期"
    ]
    ws_finance.append(finance_headers)

    finance_data = [
        ["FIN_401", "净资产", "总资产-总负债",
         "total_assets - total_liabilities",
         ">月收入*36", "需结合年龄", "季度"],
        ["FIN_402", "投资占比", "投资资产/总资产",
         "total_investments/total_assets",
         "30%-70%", "风险偏好", "月度"],
        ["FIN_403", "债务收入比", "负债/收入",
         "total_liabilities/monthly_income",
         "<50%", "预警值>80%", "月度"]
    ]
    for row in finance_data:
        ws_finance.append(row)

    # ===== 8. 行为指标 =====
    ws_behavior = wb.create_sheet("行为指标")
    behavior_headers = [
        "指标ID", "指标名称", "行为定义", "计算公式/规则",
        "基准值", "分析维度", "数据源"
    ]
    ws_behavior.append(behavior_headers)

    behavior_data = [
        ["BEH_501", "日均休闲时长", "每日娱乐活动分钟数",
         "SUM(leisure_minutes)/days",
         "120-180分钟", "年龄/职业", "行为日志"],
        ["BEH_502", "内容偏好", "最常消费内容类型",
         "MODE(content_type)",
         "N/A", "用户分群", "点击流"],
        ["BEH_503", "社交活跃度", "每周互动次数",
         "COUNT(social_actions)",
         "5-10次", "社交圈", "关系图"]
    ]
    for row in behavior_data:
        ws_behavior.append(row)

    # ===== 9. 复合指标 =====
    ws_derived = wb.create_sheet("复合指标")
    derived_headers = [
        "指标ID", "指标名称", "组成指标", "计算公式/规则",
        "业务规则", "可视化建议", "负责人"
    ]
    ws_derived.append(derived_headers)

    derived_data = [
        ["MIX_601", "健康风险指数", "BMI,血压,运动量",
         "(BMI_score*0.4)+(BP_score*0.3)+(exercise_score*0.3)",
         ">70需预警", "热力图", "医疗团队"],
        ["MIX_602", "财务健康度", "净资产,储蓄率,债务比",
         "Z-score标准化加权",
         "A/B/C分级", "雷达图", "财务部"],
        ["MIX_603", "用户价值分", "RFM,活跃度,内容贡献",
         "机器学习模型输出",
         "Top20%重点维护", "分层气泡图", "数据科学"]
    ]
    for row in derived_data:
        ws_derived.append(row)

    # ===== 10. 变更记录 =====
    ws_change = wb.create_sheet("变更记录")
    change_headers = [
        "版本", "修订日期", "修订内容", "修改人",
        "影响指标", "审批状态"
    ]
    ws_change.append(change_headers)

    change_data = [
        ["V1.0", "2025-07-01", "新增健康指标模块", "xxx",
         "HLTH_201-203", "已审批"],
        ["V1.1", "2025-07-01", "完善财务指标口径", "xxx",
         "FIN_301-303", "已审批"]
    ]
    for row in change_data:
        ws_change.append(row)

    # ===== 样式设置 =====
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet in wb:
        # 设置列宽
        for col in range(1, len(sheet[1]) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20

        # 标题行样式
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 数据区域样式
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column == 4 and sheet.title == "指标总览":  # 优先级列
                    if cell.value == "P0":
                        cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
                    elif cell.value == "P1":
                        cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")

    # ===== 保存文件 =====
    if not save_path:
        save_path = os.getcwd()

    # filename = f"用户多维指标口径模板_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filename = f"用户多维指标需求口径模板.xlsx"
    full_path = os.path.join(save_path, filename)

    # 确保目录存在
    os.makedirs(save_path, exist_ok=True)

    wb.save(full_path)
    print(f"模板已生成：{full_path}")
    return full_path


# 使用示例
create_metric_template(r"C:\Users\Licz.1\Desktop\tmp")
