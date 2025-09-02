import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_metric_template():
    # 创建工作簿
    wb = Workbook()

    # ===== 1. 封面页 =====
    ws_cover = wb.active
    ws_cover.title = "封面"
    cover_data = [
        ["指标口径需求文档"],
        ["项目名称", "经营分析"],
        ["版本号", "V1.0"],
        ["编制部门", "需求部门"],
        ["负责人", "xxx"],
        ["联系方式", "xxx"],
        ["更新日期", datetime.now().strftime("%Y-%m-%d")],
        ["说明", "此文档为核心业务指标口径定义"]
    ]
    for row in cover_data:
        ws_cover.append(row)

    # 封面样式设置
    ws_cover['A1'].font = Font(size=16, bold=True)
    for row in ws_cover.iter_rows(min_row=2, max_row=7):
        for cell in row:
            if cell.column == 1:
                cell.font = Font(bold=True)

    # ===== 2. 目录页 =====
    ws_toc = wb.create_sheet("目录")
    toc_data = [
        ["序号", "工作表名称", "内容说明", "最后更新日期"],
        [1, "指标总览", "核心指标清单", "2025-07-01"],
        [2, "基础指标", "原子级指标定义", "2025-07-01"],
        [3, "复合指标", "衍生指标计算逻辑", "2025-07-01"],
        [4, "维度管理", "分析维度及枚举值", "2025-07-01"],
        [5, "变更记录", "历史修订记录", "2025-07-01"]
    ]
    for row in toc_data:
        ws_toc.append(row)

    # 添加超链接
    for row in range(2, 7):
        ws_toc.cell(row=row, column=2).hyperlink = f"#'{ws_toc.cell(row=row, column=2).value}'!A1"

    # ===== 3. 指标总览 =====
    ws_summary = wb.create_sheet("指标总览")
    summary_headers = [
        "指标ID", "指标名称（中文）", "业务归属",
        "优先级(P0/P1/P2)", "是否核心KPI", "关联看板", "负责人"
    ]
    ws_summary.append(summary_headers)

    summary_data = [
        ["GMV_001", "成交总额", "销售部", "P0", "是", "经营日报", "xxx"],
        ["DAU_002", "日活跃用户数", "用户部", "P1", "是", "用户看板", "xxx"],
        ["CTR_003", "点击率", "运营部", "P2", "否", "活动监控", "xxx"],
        ["RET_004", "留存率", "用户部", "P1", "是", "用户看板", "xxx"]
    ]
    for row in summary_data:
        ws_summary.append(row)

    # ===== 4. 基础指标 =====
    ws_basic = wb.create_sheet("基础指标")
    basic_headers = [
        "指标ID", "指标名称", "业务定义", "计算公式/规则",
        "数据来源表", "字段名", "统计周期", "过滤条件"
    ]
    ws_basic.append(basic_headers)

    basic_data = [
        ["GMV_001", "成交总额", "已完成支付的订单总金额",
         "SUM(order_amount)", "dwd_orders", "order_amount", "日/月", "status='paid'"],
        ["DAU_002", "日活用户", "当日有操作的用户数",
         "COUNT(DISTINCT user_id)", "dwd_user_log", "user_id", "日", "event_date=当天"],
        ["UV_003", "访客数", "独立访问用户数",
         "COUNT(DISTINCT device_id)", "dwd_pv_log", "device_id", "日", "is_spider=0"],
        ["ORD_004", "订单数", "有效订单数量",
         "COUNT(order_id)", "dwd_orders", "order_id", "日", "status IN ('paid','shipped')"]
    ]
    for row in basic_data:
        ws_basic.append(row)

    # ===== 5. 复合指标 =====
    ws_derived = wb.create_sheet("复合指标")
    derived_headers = [
        "指标ID", "指标名称", "组成指标",
        "计算公式/规则", "业务规则", "可视化建议"
    ]
    ws_derived.append(derived_headers)

    derived_data = [
        ["RATE_001", "转化率", "下单用户数,访客数",
         "(下单用户数/访客数)*100%", "排除测试账号", "折线图+环比数据"],
        ["ARPU_002", "客单价", "GMV,订单数",
         "GMV/订单数", "仅计算有效订单", "柱状图+行业对比"],
        ["ROI_003", "投资回报率", "收益,成本",
         "(收益-成本)/成本*100%", "活动期间数据", "热力图+渠道维度下钻"],
        ["LTV_004", "用户生命周期价值", "总消费,用户数",
         "总消费/用户数", "按注册队列计算", "面积图+趋势分析"]
    ]
    for row in derived_data:
        ws_derived.append(row)

    # ===== 6. 维度管理 =====
    ws_dim = wb.create_sheet("维度管理")
    dim_headers = [
        "维度名称", "维度值", "业务含义",
        "技术字段名", "取值说明", "是否可下钻"
    ]
    ws_dim.append(dim_headers)

    dim_data = [
        ["渠道类型", "app", "移动端APP", "channel", "含iOS/Android", "是"],
        ["", "h5", "手机网页版", "channel", "", "是"],
        ["会员等级", "L1", "普通会员", "vip_level", "累计消费<1万元", "否"],
        ["", "L2", "白银会员", "vip_level", "1万≤消费<5万", "否"],
        ["商品类目", "电子产品", "数码家电", "category", "包含手机/电脑等", "是"],
        ["", "服装", "服饰鞋包", "category", "", "是"]
    ]
    for row in dim_data:
        ws_dim.append(row)

    # ===== 7. 变更记录 =====
    ws_change = wb.create_sheet("变更记录")
    change_headers = [
        "版本", "修订日期", "修订内容",
        "修改人", "影响范围", "审批状态"
    ]
    ws_change.append(change_headers)

    change_data = [
        ["V1.0", "2025-07-01", "初版发布", "xxx", "全部指标", "已审批"],
        ["V1.1", "2025-07-01", "新增直播相关指标", "xxx", "GMV_004-006", "已审批"],
        ["V1.2", "2025-07-01", "修正客单价计算公式", "xxx", "ARPU_002", "待审核"]
    ]
    for row in change_data:
        ws_change.append(row)

    # ===== 通用样式设置 =====
    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for sheet in wb:
        # 设置列宽
        for col in range(1, len(sheet[1]) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18

        # 标题行样式
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 数据区域样式
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.column == 4 and sheet.title == "指标总览":  # 优先级列
                    if cell.value == "P0":
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    elif cell.value == "P1":
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 指定保存路径
    custom_path = r"C:\Users\Licz.1\Desktop\tmp"
    os.makedirs(custom_path, exist_ok=True)

    filename = f"指标口径需求模板_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    full_path = os.path.join(custom_path, filename)
    print(f"文件已保存到：{full_path}")
    # 保存文件
    wb.save(full_path)
    print(f"模板已生成：{full_path}")
    return full_path


# 执行函数生成模板
create_metric_template()
