"""
export_org_hierarchy_to_excel.py

Connects to StarRocks (MySQL protocol) using STARROCKS_CONFIG from
`starrocks_utils.py`, runs the provided organization-hierarchy SQL, and writes
results into an Excel file in the same directory.

Usage:
  python export_org_hierarchy_to_excel.py [--output OUTPUT_PATH] [--no-db]

Options:
  --output OUTPUT_PATH   Path for the output Excel file (defaults to
                         ./org_hierarchy_<timestamp>.xlsx)
  --no-db                Do not connect to the database; print the SQL and
                         exit (useful for testing without credentials/network).

Dependencies: pandas, pymysql, openpyxl
"""
from __future__ import annotations
import sys
from pathlib import Path
import argparse
import datetime
import textwrap
import copy
import importlib.util
HAVE_OPENPYXL = False
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except Exception:
    openpyxl = None
    get_column_letter = None

# allow importing sibling module starrocks_utils
HERE = Path(__file__).resolve().parent
ZM_DIR = HERE.parent  # test01/zm
sys.path.insert(0, str(ZM_DIR))

try:
    import pymysql
    import pandas as pd
    # Load starrocks_utils via importlib from the local file to avoid static
    # unresolved import errors in some editors/linters and to ensure we load
    # the expected module from the repository.
    spec = importlib.util.spec_from_file_location("starrocks_utils", str(ZM_DIR / "starrocks_utils.py"))
    starrocks_utils = importlib.util.module_from_spec(spec)
    if spec and spec.loader:
        spec.loader.exec_module(starrocks_utils)
        STARROCKS_CONFIG = getattr(starrocks_utils, "STARROCKS_CONFIG", None)
    else:
        STARROCKS_CONFIG = None
except Exception as e:
    # Import errors will be raised later when running; provide friendly message
    print(f"Import warning: {e}")

SQL = textwrap.dedent("""
SELECT
    o.oid,
    o.name AS org_name,
    -- 完整层级路径
    ARRAY_JOIN([
        o1.name, o2.name, o3.name, o4.name, o5.name,
        o6.name, o7.name, o8.name, o9.name, o10.name
    ], '/') AS full_hierarchy_path,
    -- 当前组织的层级深度
    CASE
        WHEN o.tenthlevelorganization IS NOT NULL THEN 10
        WHEN o.ninthlevelorganization IS NOT NULL THEN 9
        WHEN o.eighthlevelorganization IS NOT NULL THEN 8
        WHEN o.seventhlevelorganization IS NOT NULL THEN 7
        WHEN o.sixthlevelorganization IS NOT NULL THEN 6
        WHEN o.fifthlevelorganization IS NOT NULL THEN 5
        WHEN o.fourthlevelorganization IS NOT NULL THEN 4
        WHEN o.thirdlevelorganization IS NOT NULL THEN 3
        WHEN o.secondlevelorganization IS NOT NULL THEN 2
        WHEN o.firstlevelorganization IS NOT NULL THEN 1
        ELSE 0
    END AS depth_level,
    -- 直接父级信息
    CASE
        WHEN o.tenthlevelorganization IS NOT NULL THEN CONCAT(o9.name, '(', o.ninthlevelorganization, ')')
        WHEN o.ninthlevelorganization IS NOT NULL THEN CONCAT(o8.name, '(', o.eighthlevelorganization, ')')
        WHEN o.eighthlevelorganization IS NOT NULL THEN CONCAT(o7.name, '(', o.seventhlevelorganization, ')')
        WHEN o.seventhlevelorganization IS NOT NULL THEN CONCAT(o6.name, '(', o.sixthlevelorganization, ')')
        WHEN o.sixthlevelorganization IS NOT NULL THEN CONCAT(o5.name, '(', o.fifthlevelorganization, ')')
        WHEN o.fifthlevelorganization IS NOT NULL THEN CONCAT(o4.name, '(', o.fourthlevelorganization, ')')
        WHEN o.fourthlevelorganization IS NOT NULL THEN CONCAT(o3.name, '(', o.thirdlevelorganization, ')')
        WHEN o.thirdlevelorganization IS NOT NULL THEN CONCAT(o2.name, '(', o.secondlevelorganization, ')')
        WHEN o.secondlevelorganization IS NOT NULL THEN CONCAT(o1.name, '(', o.firstlevelorganization, ')')
        ELSE '无父级'
    END AS direct_parent,
    -- 所有父级OID
    o.firstlevelorganization AS level1_parent_oid,
    o.secondlevelorganization AS level2_parent_oid,
    o.thirdlevelorganization AS level3_parent_oid,
    o.fourthlevelorganization AS level4_parent_oid,
    o.fifthlevelorganization AS level5_parent_oid,
    o.sixthlevelorganization AS level6_parent_oid,
    o.seventhlevelorganization AS level7_parent_oid,
    o.eighthlevelorganization AS level8_parent_oid,
    o.ninthlevelorganization AS level9_parent_oid,
    o.tenthlevelorganization AS level10_parent_oid,
    -- 所有父级名称
    COALESCE(o1.name, '') AS level1_parent_name,
    COALESCE(o2.name, '') AS level2_parent_name,
    COALESCE(o3.name, '') AS level3_parent_name,
    COALESCE(o4.name, '') AS level4_parent_name,
    COALESCE(o5.name, '') AS level5_parent_name,
    COALESCE(o6.name, '') AS level6_parent_name,
    COALESCE(o7.name, '') AS level7_parent_name,
    COALESCE(o8.name, '') AS level8_parent_name,
    COALESCE(o9.name, '') AS level9_parent_name,
    COALESCE(o10.name, '') AS level10_parent_name
FROM ods.ods_ehr_organization_di o
LEFT JOIN ods.ods_ehr_organization_di o1 ON o.firstlevelorganization = o1.oid AND o1.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o2 ON o.secondlevelorganization = o2.oid AND o2.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o3 ON o.thirdlevelorganization = o3.oid AND o3.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o4 ON o.fourthlevelorganization = o4.oid AND o4.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o5 ON o.fifthlevelorganization = o5.oid AND o5.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o6 ON o.sixthlevelorganization = o6.oid AND o6.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o7 ON o.seventhlevelorganization = o7.oid AND o7.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o8 ON o.eighthlevelorganization = o8.oid AND o8.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o9 ON o.ninthlevelorganization = o9.oid AND o9.stdisdeleted = 0
LEFT JOIN ods.ods_ehr_organization_di o10 ON o.tenthlevelorganization = o10.oid AND o10.stdisdeleted = 0
WHERE o.stdisdeleted = 0
ORDER BY
    -- 按层级深度排序，同层级按名称排序
    CASE
        WHEN o.tenthlevelorganization IS NOT NULL THEN 10
        WHEN o.ninthlevelorganization IS NOT NULL THEN 9
        WHEN o.eighthlevelorganization IS NOT NULL THEN 8
        WHEN o.seventhlevelorganization IS NOT NULL THEN 7
        WHEN o.sixthlevelorganization IS NOT NULL THEN 6
        WHEN o.fifthlevelorganization IS NOT NULL THEN 5
        WHEN o.fourthlevelorganization IS NOT NULL THEN 4
        WHEN o.thirdlevelorganization IS NOT NULL THEN 3
        WHEN o.secondlevelorganization IS NOT NULL THEN 2
        WHEN o.firstlevelorganization IS NOT NULL THEN 1
        ELSE 0
    END desc,
    o.name;
""")


def fetch_sql_to_dataframe(sql: str, config: dict) -> "pd.DataFrame":
    """Run SQL and return a pandas DataFrame with results.

    config: a dict compatible with pymysql.connect parameters. This function
    makes a shallow copy and removes None-valued items (like cursorclass) to
    avoid connect-time errors.
    """
    cfg = copy.copy(config)
    # ensure port is int
    if 'port' in cfg:
        try:
            cfg['port'] = int(cfg['port'])
        except Exception:
            pass
    # remove None values
    cfg = {k: v for k, v in cfg.items() if v is not None}

    conn = pymysql.connect(**cfg)
    try:
        with conn.cursor() as cur:
            cur.execute(sql)
            rows = cur.fetchall()
            columns = [d[0] for d in cur.description] if cur.description else []
    finally:
        conn.close()

    df = pd.DataFrame(rows, columns=columns)
    return df


def main() -> int:
    p = argparse.ArgumentParser(description="Export organization hierarchy to Excel from StarRocks")
    p.add_argument('--output', '-o', help='Output Excel path')
    p.add_argument('--no-db', action='store_true', help='Do not connect to DB; print SQL and exit')
    args = p.parse_args()

    if args.no_db:
        print("--no-db specified. SQL below:\n")
        print(SQL)
        return 0

    out_path = Path(args.output) if args.output else (HERE / f"org_hierarchy_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    try:
        df = fetch_sql_to_dataframe(SQL, STARROCKS_CONFIG)
    except Exception as e:
        print(f"Failed to fetch data from StarRocks: {e}")
        return 2

    # Map DataFrame column names to Chinese headers for the Excel file.
    HEADER_MAP = {
        'oid': '组织OID',
        'org_name': '组织名称',
        'full_hierarchy_path': '完整层级路径',
        'depth_level': '层级深度',
        'direct_parent': '直接父级',
        'level1_parent_oid': '父级1_OID',
        'level2_parent_oid': '父级2_OID',
        'level3_parent_oid': '父级3_OID',
        'level4_parent_oid': '父级4_OID',
        'level5_parent_oid': '父级5_OID',
        'level6_parent_oid': '父级6_OID',
        'level7_parent_oid': '父级7_OID',
        'level8_parent_oid': '父级8_OID',
        'level9_parent_oid': '父级9_OID',
        'level10_parent_oid': '父级10_OID',
        'level1_parent_name': '父级1_名称',
        'level2_parent_name': '父级2_名称',
        'level3_parent_name': '父级3_名称',
        'level4_parent_name': '父级4_名称',
        'level5_parent_name': '父级5_名称',
        'level6_parent_name': '父级6_名称',
        'level7_parent_name': '父级7_名称',
        'level8_parent_name': '父级8_名称',
        'level9_parent_name': '父级9_名称',
        'level10_parent_name': '父级10_名称',
    }

    # Only rename columns that exist in the fetched DataFrame to avoid KeyError.
    existing_map = {k: v for k, v in HEADER_MAP.items() if k in df.columns}
    if existing_map:
        df = df.rename(columns=existing_map)

    try:
        df.to_excel(out_path, index=False)
    except Exception as e:
        print(f"Failed to write Excel file: {e}")
        return 3

    # Post-process the Excel file with openpyxl: freeze header row and auto-adjust column widths
    if HAVE_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(out_path)
            ws = wb.active
            # Freeze header row
            ws.freeze_panes = 'A2'

            # Auto-adjust column widths based on content
            for i, column_cells in enumerate(ws.columns, 1):
                col_letter = get_column_letter(i)
                max_length = 0
                for cell in column_cells:
                    try:
                        cell_value = '' if cell.value is None else str(cell.value)
                    except Exception:
                        cell_value = ''
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(out_path)
        except Exception as e:
            print(f"Warning: wrote Excel but failed post-processing with openpyxl: {e}")
            # Not a fatal error; the Excel file was already written.
    else:
        print("Note: openpyxl not installed; skipped Excel post-processing (freeze/auto-width). Install openpyxl to enable this.")

    print(f"Exported {len(df)} rows to {out_path}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main())

